#!/usr/bin/env python3
# Excelify2.py – Create a styled Excel workbook from the fully populated plant CSV.
# 2025-06-13 · Adds blanket COLUMN_WIDTHS dict, keep_default_na, and cleans up width logic.
# todo I need to make the first sheet the "pretty sheet that you can interface with, There needs to be an appended "RAW CSV" tab the houses the data we are going to export.
# todo I need to make the "RAW CSV" editable/viewable from the first page
# todo "Link: Others" needs to be autopopulated on the first sheet split up into columns for each entry:
# todo CSV Format: [Tag1,"URL1","Label1"];[Tag2,"URL2","Label2"] : "[T1,""https://Test.com"",""Test 1""];[T2,""https://Test.com"",""Test 2""];[T3,""https://Test.com"",""Test 3""]"
# todo 
# todo Excel Format   -- Maybe a Searchable/appendable list
# todo   |    aa    |     ab     |      ac    |   
# todo   |          Other : Links             |
# todo   |  Label1  |    URL1    |    Tag1    |
# todo   |  Label2  |    URL2    |    Tag2    |
# ? I need to figure out how to get urls added and edited from the CSV header


from pathlib import Path
import sys, argparse, pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
import black

# ── Column widths (characters) -------------------------------------------
# Tweak values here to resize any column on export.
COLUMN_WIDTHS: dict[str, int] = {
    # core ID fields
    "Plant Type": 20,
    "Key": 7,
    "Botanical Name": 23,
    "Common Name": 23,
    # basics
    "Height (ft)": 10,
    "Spread (ft)": 10,
    "Bloom Color": 18,
    "Bloom Time": 18,
    "Sun": 18,
    "Water": 15,
    "AGCP Regional Status": 18,
    "USDA Hardiness Zone": 16,
    # descriptive
    "Attracts": 24,
    "Tolerates": 28,
    "Soil Description": 32,
    "Condition Comments": 32,
    "MaintenanceLevel": 15,
    "Native Habitats": 26,
    "Culture": 24,
    "Uses": 20,
    "UseXYZ": 20,
    "WFMaintenance": 22,
    "Problems": 25,
    # links
    "Link: Missouri Botanical Garden": 27,
    "Link: Wildflower.org": 27,
    "Link: Pleasantrunnursery.com": 27,
    "Link: Newmoonnursery.com": 27,
    "Link: Pinelandsnursery.com": 27,
    # review
    "Rev": 15,
    "Mark Reviewed": 8,
}
DEFAULT_WIDTH: int = 18

# ── CLI ------------------------------------------------------------------
parser = argparse.ArgumentParser(description="Export formatted Excel from CSV")
parser.add_argument("--in_csv",      default="Outputs/Plants_Linked_Filled.csv")
parser.add_argument("--out_xlsx",    default="Outputs/Plants_Linked_Filled_NeedsReview.xlsx")
parser.add_argument("--template_csv",default="Templates/Plants_Template.csv")
args = parser.parse_args()

# ── Paths ----------------------------------------------------------------
def repo_dir() -> Path:
    exe_path = Path(sys.executable if getattr(sys, "frozen", False) else __file__).resolve()
    for parent in exe_path.parents:
        if (parent / "Templates").is_dir() and (parent / "Outputs").is_dir():
            return parent
    return exe_path.parent

REPO         = repo_dir()
CSV_FILE     = (REPO / args.in_csv).resolve()
XLSX_FILE    = (REPO / args.out_xlsx).resolve()
TEMPLATE_CSV = (REPO / args.template_csv).resolve()
XLSX_FILE.parent.mkdir(parents=True, exist_ok=True)

# ── Load CSV, align columns ----------------------------------------------
df = (
    pd.read_csv(
        CSV_FILE,
        dtype=str,
        encoding="utf-8-sig",
        keep_default_na=False,       # ← keeps literal “NA” strings
    )
    .fillna("")
)
template_cols = list(pd.read_csv(TEMPLATE_CSV, nrows=0, keep_default_na=False).columns)
df = df.reindex(columns=template_cols + [c for c in df.columns if c not in template_cols])

# Normalise casing
if "Common Name" in df.columns:
    df["Common Name"] = df["Common Name"].str.upper()
if "Botanical Name" in df.columns:
    df["Botanical Name"] = df["Botanical Name"].apply(
        lambda n: " ".join(w.capitalize() for w in str(n).split())
    )

# ── Inject “Mark Reviewed” column if missing -----------------------------
if "Mark Reviewed" not in df.columns:
    df.insert(df.columns.get_loc("Rev") + 1, "Mark Reviewed", "")

df.to_excel(XLSX_FILE, index=False, na_rep="NA")
wb = load_workbook(XLSX_FILE)
ws = wb.active
ws.title = "Plant Data"

# ── Source-legend row (unchanged) ----------------------------------------
DATA_SOURCE = {
    "Plant Type": "Masterlist",
    "Key": "FillMissingData",
    "Botanical Name": "Masterlist",
    "Common Name": "Masterlist",
    "Height (ft)": "MBG -> WF -> Pinelands",
    "Spread (ft)": "MBG -> WF -> Pinelands",
    "Bloom Color": "WF + MBG + Pinelands/NM",
    "Bloom Time": "WF + MBG + Pinelands/NM",
    "Sun": "MBG -> WF “Light Req.”",
    "Water": "MBG -> WF “Soil Moisture”",
    "AGCP Regional Status": "WF (Wetland Indicator)",
    "USDA Hardiness Zone": "MBG “Zone”",
    "Attracts": "PR + WF + MBG + Pinelands",
    "Tolerates": "MBG + PR + NM + Pinelands",
    "Soil Description": "WF “Soil Description”",
    "Condition Comments": "WF “Condition Comments”",
    "MaintenanceLevel": "MBG “Maintenance”",
    "Native Habitats": "WF “Native Habitat”",
    "Culture": "MBG “Culture”",
    "Uses": "MBG “Uses”",
    "UseXYZ": "WF Benefit list",
    "WFMaintenance": "WF “Maintenance:”",
    "Problems": "MBG “Problems”",
    "Link: Missouri Botanical Garden": "GetLinks (MBG ID)",
    "Link: Wildflower.org": "GetLinks (USDA ID)",
    "Link: Pleasantrunnursery.com": "GetLinks (name)",
    "Link: Newmoonnursery.com": "GetLinks (name)",
    "Link: Pinelandsnursery.com": "GetLinks (name)",
    "Rev": "User Input (YYYYMMDD_FL)",
    "Mark Reviewed": "Type Initials; Inserts YYYYMMDD_FL",
}

ws.insert_rows(2)
for col_idx, col_name in enumerate([c.value for c in ws[1]], start=1):
    cell = ws.cell(row=2, column=col_idx)
    cell.value = DATA_SOURCE.get(col_name, "")
    cell.font = Font(italic=True, size=8)
    cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=False, shrink_to_fit=False)

# ── Header formatting + freeze -------------------------------------------
HEADER_FILL = PatternFill(start_color="CFE2F3", end_color="CFE2F3", fill_type="solid")
for c in ws[1]:
    c.fill = HEADER_FILL
    c.font = Font(bold=True, size=11)
    c.alignment = Alignment(horizontal="center", vertical="center")
ws.freeze_panes = "E3"

# ── Autofit helper --------------------------------------------------------
def autofit(ws: Worksheet) -> None:
    long_fields = {
        "UseXYZ": 50, "Culture": 50, "Uses": 50,
        "Soil Description": 48, "Condition Comments": 48, "Native Habitats": 42,
    }
    headers = [c.value for c in ws[1]]
    for col_cells in ws.iter_cols(min_row=1, max_row=ws.max_row):
        col_idx = col_cells[0].column
        col_letter = get_column_letter(col_idx)
        header = headers[col_idx - 1] if col_idx - 1 < len(headers) else ""
        lengths = [
            len(str(cell.value or ""))
            for cell in col_cells
            if str(cell.value or "").strip() not in {"Needs Review", ""}
        ]
        if lengths:
            cap = long_fields.get(header, 64)
            ws.column_dimensions[col_letter].width = min(max(lengths) + 2, cap)

# ── Style all data cells (unchanged) -------------------------------------
MISSING_FILL = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
NA_LINK_FILL = PatternFill(start_color="B7D7FF", end_color="B7D7FF", fill_type="solid")
REV_FILLED   = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

def style_sheet(ws: Worksheet, df: pd.DataFrame, header: list[str]) -> None:
    REV_MISSING_FILL = PatternFill(start_color="FFF79A", end_color="FFF79A", fill_type="solid")
    ROW_ALT_FILL = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")
    long_wrap_fields = {
        # already here:
        "UseXYZ", "Culture", "Uses",
        "Soil Description", "Condition Comments", "Native Habitats",
        "Rev",
        # add the two list-heavy ones:
        "Attracts", "Tolerates",
    }


    for r_idx, row in enumerate(df.itertuples(index=False, name=None), start=3):
        for c_idx, (col_name, val) in enumerate(zip(header, row), start=1):
            cell = ws.cell(row=r_idx, column=c_idx)
            val = str(val).strip()
            col_name_lower = col_name.strip().lower()

            # Rev column handling
            if col_name_lower == "rev":
                if val:
                    cell.value = val
                    cell.fill = REV_FILLED
                else:
                    mark_col_letter = get_column_letter(c_idx + 1)
                    cell.value = (
                        f'=IF({mark_col_letter}{r_idx}<>"",'
                        f' TEXT(TODAY(),"yyyymmdd") & "_" & {mark_col_letter}{r_idx}, "")'
                    )
                    cell.fill = REV_MISSING_FILL
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                continue

            # Mark Reviewed column
            if col_name_lower == "mark reviewed":
                cell.value = val if val else ""
                cell.alignment = Alignment(horizontal="center", vertical="center")
                continue

            # Links – NA / Needs Review logic
            if val.upper() == "NA" and col_name.startswith("Link: "):
                cell.value, cell.fill = "NA", NA_LINK_FILL
            elif not val:
                cell.value, cell.fill = "Needs Review", MISSING_FILL
            else:
                cell.value = val
                if col_name.startswith("Link: ") and val.lower().startswith("http"):
                    cell.hyperlink = val
                    cell.style = "Hyperlink"
                    cell.font = Font(color="0000EE", underline="single")

                # Italicise Botanical Name
                if col_name_lower == "botanical name":
                    parts = val.split()
                    if len(parts) >= 2:
                        genus, species = parts[0].capitalize(), parts[1].lower()
                        variety = " ".join(parts[2:]).strip("' ")
                        variety = f"'{variety.title()}'" if variety else ""
                        cell.value = " ".join(filter(None, [genus, species, variety]))
                    cell.font = Font(italic=True)

            # Text wrapping
            if col_name in long_wrap_fields:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            else:
                cell.alignment = Alignment(wrap_text=False, vertical="top")

        # alt-row shading
        if r_idx % 2 == 1:
            for cur in ws[r_idx]:
                if cur.fill == PatternFill():
                    cur.fill = ROW_ALT_FILL
        
                # ── auto-grow row height when wrapping is on --------------------
        max_lines = 1
        for cur in ws[r_idx]:
            if cur.alignment and cur.alignment.wrap_text:
                # a) explicit line-breaks             → count '\n'
                # b) long text that Excel will wrap   → rough 50-char estimate
                lines = max(
                    str(cur.value).count("\n") + 1,
                    len(str(cur.value or "")) // 50 + 1,
                )
                max_lines = max(max_lines, lines)

        if max_lines > 1:
            ws.row_dimensions[r_idx].height = 15 * max_lines  # ≈15 px per line


# ── Style + Fit ----------------------------------------------------------
style_sheet(ws, df, df.columns.tolist())
autofit(ws)

for hdr_cell in ws[1]:
    hdr = str(hdr_cell.value).strip()
    col_letter = get_column_letter(hdr_cell.column)
    ws.column_dimensions[col_letter].width = COLUMN_WIDTHS.get(hdr, DEFAULT_WIDTH)

# Row heights
for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
    ws.row_dimensions[row[0].row].height = 28

# ── README -------------------------------------------------------------------
readme = wb.create_sheet("README")
readme.sheet_properties.tabColor = "A9A9A9"
readme.column_dimensions["A"].width = 100

readme["A1"] = "Instructions:"
readme["A3"] = "To record a review:"
readme["A4"] = "1. Type your initials (e.g., AN) in the 'Mark Reviewed' column."
readme["A5"] = "2. The 'Rev' column will auto-fill with today's date and your initials."
readme["A7"] = "Legend:"
readme["A8"] = "RED: Missing value (empty cell)"
readme["A9"] = "BLUE: Link doesn't exist and was reviewd. ‘NA’ =/= No Value"

readme["A11"] = "Filters applied only to these columns:"
readme["A12"] = ", ".join(["Plant Type", "Bloom Color", "Sun", "Water", "Attracts"])

readme["A14"] = "Tools: How to filter by partial match in Excel:"
readme["A15"] = "1. Click the filter dropdown on the header."
readme["A16"] = "2. Choose 'Text Filters' ➜ 'Contains...'"
readme["A17"] = "3. Type part of a word (e.g., 'shade', 'yellow')."


# -- Step 6 · embed helper scripts ----------------------------------------
from pathlib import Path
import black

def find_script_root(repo: Path) -> Path:
    """
    Locate the folder that holds the full-version helper scripts.

    1) <repo>/Static/Python_full           ← current layout
    2) <repo>/_internal/Static/Python_full ← legacy layout
    """
    for candidate in (
        repo / "Static" / "Python_full",
        repo / "_internal" / "Static" / "Python_full",
    ):
        if candidate.is_dir():
            return candidate.resolve()
    raise FileNotFoundError("Cannot locate Static/Python_full")

PYTHON_FULL = find_script_root(REPO)

script_descriptions = {
    "PDFScraper.py":      "Extract plant data from source PDF",
    "GetLinks.py":        "Find MBG & WF links",
    "FillMissingData.py": "Populate missing fields",
    "GeneratePDF.py":     "Create formatted PDF guide",
    "Excelify2.py":       "Make this Excel workbook",
}

for script, desc in script_descriptions.items():
    src = PYTHON_FULL / script
    if not src.exists():
        continue                       # skip helpers that are not present
    with src.open(encoding="utf-8") as f:
        raw = f.read()
    try:
        code = black.format_str(raw, mode=black.Mode())
    except Exception:
        code = raw                     # keep original text if Black fails
    ws_code = wb.create_sheet(script)  # one worksheet per helper script
    ws_code.column_dimensions["A"].width = 120
    ws_code["A1"] = f"# {script} - {desc}"
    for i, line in enumerate(code.splitlines(), start=2):
        ws_code[f"A{i}"] = line


# ── Step 7 · pip requirements list -----------------------------------------
req = REPO / "requirements.txt"
row = readme.max_row + 2
readme[f"A{row}"] = "[PKG] Required Python Packages:"
if req.exists():
    for i, line in enumerate(req.read_text().splitlines(), start=row+1):
        if line.strip() and not line.lstrip().startswith("#"):
            readme[f"A{i}"] = line.strip()

# ── Step 8 · pull full README.md (optional) ---------------------------------
readme_md = REPO / "readme.md"
if readme_md.exists():
    tab = wb.create_sheet("README_full")
    tab.column_dimensions["A"].width = 120
    for i, line in enumerate(readme_md.read_text(encoding="utf-8").splitlines(), start=1):
        tab[f"A{i}"] = line

# ── Step 9 · finish ---------------------------------------------------------
wb.save(XLSX_FILE)
print(f"Yeehaw! Workbook saved -> {XLSX_FILE}")
