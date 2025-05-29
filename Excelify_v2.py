# /plants/Excelify_v2.py
# Enhanced version with:
# - Stats and Summary tabs
# - Improved missing value highlighting
# - Column filters applied selectively
# - Versioned script overview in README

# /plants/Excelify.py

from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

BASE      = Path(__file__).resolve().parent
CSV_FILE  = BASE / "Plants_COMPLETE.csv"
XLSX_FILE = CSV_FILE.with_suffix(".xlsx")

# ─── Load DataFrame and write to Excel ────────────────────────────────────
df = pd.read_csv(CSV_FILE, dtype=str).fillna("")
df.to_excel(XLSX_FILE, index=False)
wb = load_workbook(XLSX_FILE)
ws = wb.active
ws.title = "Plant Data"

# ─── Freeze Panes, Style Headers, Auto Width ──────────────────────────────
ws.freeze_panes = "A2"
HEADER_FILL = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
BOLD_FONT   = Font(bold=True)
for cell in ws[1]:
    cell.fill = HEADER_FILL
    cell.font = BOLD_FONT
for i, column_cells in enumerate(ws.columns, 1):
    max_length = max(len(str(cell.value or "")) for cell in column_cells)
    ws.column_dimensions[get_column_letter(i)].width = min(max_length + 2, 50)

# ─── Filter only specified columns ────────────────────────────────────────
filter_cols = ["Page in PDF", "Plant Type", "Bloom Color", "Sun", "Water", "Characteristics"]
header = [cell.value for cell in ws[1]]
filter_indices = [i+1 for i, val in enumerate(header) if val in filter_cols]
if filter_indices:
    col_range = f"{get_column_letter(min(filter_indices))}1:{get_column_letter(max(filter_indices))}1"
    ws.auto_filter.ref = col_range

# ─── Cell Highlighting: Only Missing Values ───────────────────────────────
RED = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
for row_idx, row in enumerate(df.itertuples(index=False), start=2):
    for col_idx, col_name in enumerate(header, start=1):
        cell = ws.cell(row=row_idx, column=col_idx)
        if not cell.value or str(cell.value).strip() == "":
            cell.fill = RED

# ─── README Sheet with Instructions ───────────────────────────────────────
readme = wb.create_sheet("README")
readme.sheet_properties.tabColor = "A9A9A9"
readme.column_dimensions["A"].width = 100

readme["A1"] = "🌿 Plant Database Export: Excel Legend and Process Notes"
readme["A3"] = "Legend:"
readme["A4"] = "🔴 Red: Missing value (empty cell)"

readme["A6"]  = "Filters applied only to these columns:"
readme["A7"]  = "Page in PDF, Plant Type, Bloom Color, Sun, Water, Characteristics"

readme["A9"]  = "🛠 How to filter by partial match in Excel:"
readme["A10"] = "1. Click the filter dropdown on the column header (e.g., Sun or Characteristics)."
readme["A11"] = "2. Choose 'Text Filters' > 'Contains...'"
readme["A12"] = "3. Type a partial term (e.g., 'shade', 'yellow') and click OK."
readme["A13"] = "💡 You can use this to find plants matching conditions across categories."
readme["A15"] = "📄 https://github.com/InfusedChooch/Plants."

# ─── Add Script Version Info ──────────────────────────────────────────────
script_descriptions = {
    "PDFScrape.py"          : "Extracts plant data from the PDF guide",
    "GetLinks_FullHybrid.py": "Finds official MBG and WF URLs for each plant",
    "fill.py"               : "Populates missing fields using those links",
    "TestLinks.py"          : "Validates that all stored links return a live page",
    "Excelify.py"           : "Creates formatted Excel output with filters and highlights",
}
row_start = readme.max_row + 2
readme[f"A{row_start}"] = "📁 Script Version Info (Last Modified):"
for i, (filename, description) in enumerate(script_descriptions.items(), start=row_start + 1):
    path = BASE / filename
    if path.exists():
        modified = datetime.fromtimestamp(path.stat().st_mtime).strftime("%Y-%m-%d %H:%M")
        readme[f"A{i}"] = f"{filename:<24} → {modified}    {description}"
    else:
        readme[f"A{i}"] = f"{filename:<24} → MISSING       {description}"

# ─── Append pip requirements to README (from static file) ────────────────
req_path = BASE / "requirements.txt"
readme_row = readme.max_row + 2
readme[f"A{readme_row}"] = "📦 Required Python Packages:"
try:
    lines = req_path.read_text(encoding="utf-8").splitlines()
    for i, line in enumerate(lines, start=readme_row + 1):
        if line.strip() and not line.strip().startswith("#"):
            readme[f"A{i}"] = line.strip()
except Exception as e:
    readme[f"A{readme_row + 1}"] = f"⚠️ Error reading requirements.txt: {e}"


# ─── Individual Markdown-Formatted Code Sheets ────────────────────────────
for script_name in script_descriptions:
    path = BASE / script_name
    if not path.exists():
        continue
    ws = wb.create_sheet(script_name)
    ws.column_dimensions["A"].width = 120
    ws["A1"] = "```python"
    with open(path, "r", encoding="utf-8") as f:
        for i, line in enumerate(f, start=2):
            ws[f"A{i}"] = line.rstrip("\n")
    ws[f"A{i+1}"] = "```"

# ─── Save Excel File ──────────────────────────────────────────────────────
wb.save(XLSX_FILE)
print(f"📘 Final Excel saved → {XLSX_FILE}")
