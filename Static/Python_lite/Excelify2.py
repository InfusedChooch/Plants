#!/usr/bin/env python3
# Excelify2.py - Create a styled Excel workbook from the fully populated plant CSV.

from pathlib import Path
import sys, argparse, pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from datetime import datetime
import black


# --- CLI ------------------------------------------------------------------
parser = argparse.ArgumentParser(description="Export formatted Excel from CSV")
parser.add_argument(
    "--in_csv",
    default="Outputs/Plants_Linked_Filled.csv",  # <- moved
    help="Input CSV file with filled data",
)
parser.add_argument(
    "--out_xlsx",
    default="Outputs/Plants_Linked_Filled_Review.xlsx",  # <- moved
    help="Output Excel file",
)
parser.add_argument(
    "--template_csv",
    default="Templates/Plants_Linked_Filled_Master.csv",  # <- moved
    help="CSV file containing column template",
)
args = parser.parse_args()


# --- Path helpers ---------------------------------------------------------
def repo_dir() -> Path:
    """Folder that contains the helper EXE (frozen) or repo root (source)."""
    if getattr(sys, "frozen", False):
        exe_dir = Path(sys.executable).resolve().parent
        return exe_dir.parent if exe_dir.name.lower() == "helpers" else exe_dir
    # Python scripts live two levels below the repo root
    return Path(__file__).resolve().parent.parent


REPO = repo_dir()
CSV_FILE = (REPO / args.in_csv).resolve()
XLSX_FILE = (REPO / args.out_xlsx).resolve()
TEMPLATE_CSV = (REPO / args.template_csv).resolve()

# ensure the Outputs folder exists when running on a flash-drive
XLSX_FILE.parent.mkdir(parents=True, exist_ok=True)

# --- rest of the script stays unchanged -----------------------------------


# --- Step 1: Load CSV and write it to a basic Excel file ------------------
df = pd.read_csv(CSV_FILE, dtype=str).fillna("")
template_cols = list(pd.read_csv(TEMPLATE_CSV, nrows=0).columns)
df = df.reindex(
    columns=template_cols + [c for c in df.columns if c not in template_cols]
)

# Format Common and Botanical Name
if "Common Name" in df.columns:
    df["Common Name"] = df["Common Name"].str.upper()

if "Botanical Name" in df.columns:
    df["Botanical Name"] = df["Botanical Name"].apply(
        lambda name: " ".join(w.capitalize() for w in str(name).split())
    )

df.to_excel(XLSX_FILE, index=False)
wb = load_workbook(XLSX_FILE)
ws = wb.active
ws.title = "Plant Data"

# --- Step 2: Style Headers and Columns ------------------------------------
ws.freeze_panes = "A2"
HEADER_FILL = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
BOLD_FONT = Font(bold=True)

for cell in ws[1]:
    cell.fill = HEADER_FILL
    cell.font = BOLD_FONT


def autofit_columns(ws: Worksheet) -> None:
    """Resize each column based solely on its header label."""

    for cell in ws[1]:
        letter = get_column_letter(cell.col_idx)
        ws.column_dimensions[letter].width = len(str(cell.value or "")) + 2


autofit_columns(ws)

# --- Step 3: Apply Filters ------------------------------------------------
filter_cols = ["Plant Type", "Bloom Color", "Sun", "Water", "Attracts"]
header = [cell.value for cell in ws[1]]
filter_indices = [i + 1 for i, val in enumerate(header) if val in filter_cols]
if filter_indices:
    col_range = f"{get_column_letter(min(filter_indices))}1:{get_column_letter(max(filter_indices))}1"
    ws.auto_filter.ref = col_range

# --- Step 4: Format Cells + Short Hyperlinks ------------------------------
link_map = {
    "Link: Missouri Botanical Garden": "[MBG]",
    "Link: Wildflower.org": "[WF]",
    "Link: Pleasantrunnursery.com": "[PR]",
    "Link: Newmoonnursery.com": "[NM]",
    "Link: Pinelandsnursery.com": "[PN]",
}


def style_sheet(ws: Worksheet, df: pd.DataFrame, header: list[str]) -> None:
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

    for row_idx, row in enumerate(df.itertuples(index=False, name=None), start=2):
        for col_idx, (col_name, value) in enumerate(zip(header, row), start=1):
            cell: Cell = ws.cell(row=row_idx, column=col_idx)
            value = str(value).strip()

            if col_name in link_map and value.startswith("http"):
                cell.value = link_map[col_name]
                cell.hyperlink = value
                cell.style = "Hyperlink"
            else:
                cell.value = value

            if col_name == "Botanical Name":
                cell.font = Font(italic=True)
            if not value:
                cell.fill = red_fill


style_sheet(ws, df, header)

# --- Step 4B: Add raw export sheet with full link data ---------------------
raw_sheet = wb.create_sheet("Plant Data CSV - No Short Links")
for col_idx, col_name in enumerate(df.columns, start=1):
    cell = raw_sheet.cell(row=1, column=col_idx)
    cell.value = col_name
    cell.font = BOLD_FONT

for row_idx, row in enumerate(df.itertuples(index=False, name=None), start=2):
    for col_idx, value in enumerate(row, start=1):
        raw_sheet.cell(row=row_idx, column=col_idx).value = value

# --- Step 4C: Set Column Widths -------------------------------------------
excel_widths = {
    "A": 22,
    "B": 9,
    "C": 26,
    "D": 23,
    "E": 13,
    "G": 13,
    "H": 22,
    "I": 22,
    "J": 15,
    "K": 18,
    "L": 18,
    "M": 43,
    "N": 26,
    "O": 25,
    "P": 25,
    "Q": 22,
    "R": 15,
    "S": 15,
    "T": 15,
    "U": 15,
    "V": 15,
    "W": 15,
}

for sheet in (ws, raw_sheet):
    for letter, width in excel_widths.items():
        sheet.column_dimensions[letter].width = width

# --- Step 5: README Sheet -------------------------------------------------
readme = wb.create_sheet("README")
readme.sheet_properties.tabColor = "A9A9A9"
readme.column_dimensions["A"].width = 100
readme["A1"] = " Plant Database Export: Excel Legend and Process Notes"
readme["A3"] = "Legend:"
readme["A4"] = "RED: Red: Missing value (empty cell)"
readme["A6"] = "Filters applied only to these columns:"
readme["A7"] = ", ".join(filter_cols)
readme["A9"] = "Tools: How to filter by partial match in Excel:"
readme["A10"] = (
    "1. Click the filter dropdown on the column header (e.g., Sun or Characteristics)."
)
readme["A11"] = "2. Choose 'Text Filters' > 'Contains...'"
readme["A12"] = "3. Type a partial term (e.g., 'shade', 'yellow') and click OK."
readme["A13"] = "Tip: Use this to find plants matching conditions across categories."
readme["A15"] = "Info: https://github.com/InfusedChooch/Plants"
readme["A16"] = (
    "This Excel was generated from the filled CSV using the script Excelify2.py"
)
readme["A17"] = (
    "Download https://portableapps.com/apps/internet/google_chrome_portable and place it in the /Static folder"
)
readme["A18"] = "Static/GoogleChromePortable"

# --- Step 6: Script Version Info ------------------------------------------
script_descriptions = {
    "Static/Python_full/PDFScraper.py": "Extracts plant data from the PDF guide",
    "Static/Python_full/GetLinks.py": "Finds official MBG & WF URLs for each plant",
    "Static/Python_full/FillMissingData.py": "Populates missing fields using those links",
    "Static/Python_full/GeneratePDF.py": "Creates printable PDF guide with images and sections",
    "Static/Python_full/Excelify2.py": "Creates formatted Excel output with filters & highlights",
}
row_start = readme.max_row + 2
readme[f"A{row_start}"] = "Folder: Script Version Info (Last Modified):"
for i, (script_path, description) in enumerate(
    script_descriptions.items(), start=row_start + 1
):
    full_path = REPO / script_path
    if full_path.exists():
        modified = datetime.fromtimestamp(full_path.stat().st_mtime).strftime(
            "%Y-%m-%d %H:%M"
        )
        readme[f"A{i}"] = f"{script_path:<40} -> {modified}    {description}"
    else:
        readme[f"A{i}"] = f"{script_path:<40} -> MISSING        {description}"

# --- Step 7: Add pip requirements -----------------------------------------
req_path = REPO / "requirements.txt"
readme_row = readme.max_row + 2
readme[f"A{readme_row}"] = "[PKG] Required Python Packages:"
try:
    lines = req_path.read_text(encoding="utf-8").splitlines()
    for i, line in enumerate(lines, start=readme_row + 1):
        if line.strip() and not line.strip().startswith("#"):
            readme[f"A{i}"] = line.strip()
except Exception as e:
    readme[f"A{readme_row + 1}"] = f"[WARN] Error reading requirements.txt: {e}"

# --- Step 8: Embed Black-formatted code ----------------------------------
for script_path, description in script_descriptions.items():
    full_path = REPO / script_path
    if not full_path.exists():
        continue

    with open(full_path, "r", encoding="utf-8") as f:
        raw_code = f.read()

    try:
        formatted_code = black.format_str(raw_code, mode=black.Mode())
    except Exception:
        formatted_code = raw_code

    ws_embed = wb.create_sheet(Path(script_path).name)
    ws_embed.column_dimensions["A"].width = 120
    ws_embed["A1"] = "```python"
    for i, line in enumerate(formatted_code.splitlines(), start=2):
        ws_embed[f"A{i}"] = line
    ws_embed[f"A{i+1}"] = "```"

# --- Step 9: Import README.md ---------------------------------------------
readme_md_path = REPO / "readme.md"
if readme_md_path.exists():
    readme_full = wb.create_sheet("README_full")
    readme_full.column_dimensions["A"].width = 120
    with readme_md_path.open("r", encoding="utf-8") as f:
        for i, line in enumerate(f, start=1):
            readme_full[f"A{i}"] = line.rstrip("\n")
else:
    print("[WARN] readme.md not found. Skipping README_full tab.")

# --- Step 10: Save and Done ----------------------------------------------
wb.save(XLSX_FILE)
print(f"Yeehaw Final Excel saved --> {XLSX_FILE}")
