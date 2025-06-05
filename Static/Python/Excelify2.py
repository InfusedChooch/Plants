# Static\Python\Excelify2.py
# Description: Create a styled Excel workbook from the fully populated plant CSV.

from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from datetime import datetime
import black
import argparse
import sys

parser = argparse.ArgumentParser(description="Export formatted Excel from CSV")
parser.add_argument(
    "--in_csv",
    default="Static/Outputs/Plants_Linked_Filled.csv",
    help="Input CSV file with filled data",
)
parser.add_argument(
    "--out_xlsx",
    default="Static/Outputs/Plants_Linked_Filled_Review.xlsx",
    help="Output Excel file",
)
args = parser.parse_args()

# ─── File Paths ───────────────────────────────────────────────────────────
BASE = Path(__file__).resolve().parent
REPO = BASE.parent.parent
CSV_FILE = (REPO / args.in_csv).resolve()
XLSX_FILE = (REPO / args.out_xlsx).resolve()

# ─── Step 1: Load CSV and write it to a basic Excel file ──────────────────
df = pd.read_csv(CSV_FILE, dtype=str).fillna("")
template_cols = list(
    pd.read_csv(
        REPO / "Static/Templates/Plants_Linked_Filled_Master.csv", nrows=0
    ).columns
)
df = df.reindex(
    columns=template_cols + [c for c in df.columns if c not in template_cols]
)

# Format Common and Botanical Name
if "Common Name" in df.columns:
    df["Common Name"] = df["Common Name"].str.upper()

if "Botanical Name" in df.columns:
    df["Botanical Name"] = df["Botanical Name"].apply(
        lambda name: " ".join(w.capitalize() for w in str(name).split())
    )

df.to_excel(XLSX_FILE, index=False)
wb = load_workbook(XLSX_FILE)
ws = wb.active
ws.title = "Plant Data"

# ─── Step 2: Style Headers and Columns ────────────────────────────────────
# ─── Step 2: Style Headers and Columns ────────────────────────────────────
ws.freeze_panes = "A2"
HEADER_FILL = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
BOLD_FONT = Font(bold=True)

for cell in ws[1]:
    cell.fill = HEADER_FILL
    cell.font = BOLD_FONT


def autofit_columns(ws: Worksheet) -> None:
    """Resize each column based solely on its header label."""

    for cell in ws[1]:
        letter = get_column_letter(cell.col_idx)
        ws.column_dimensions[letter].width = len(str(cell.value or "")) + 2


def set_fixed_column_widths(ws: Worksheet) -> None:
    """Apply hard-coded column widths using Excel character units."""

    pixel_widths = {
        "A": 150.0,
        "B": 60.0,
        "C": 180.0,
        "D": 160.0,
        "E": 90.0,
        "G": 90.0,
        "H": 150.0,
        "I": 150.0,
        "J": 100.0,
        "K": 125.0,
        "L": 120.0,
        "M": 300.0,
        "N": 180.0,
        "O": 175.0,
        "P": 175.0,
        "Q": 150.0,
        "R": 100.0,
        "S": 100.0,
        "T": 100.0,
        "U": 100.0,
        "V": 100.0,
        "W": 100.0,
    }

    # Excel stores widths as the number of "0" characters that fit in the column.
    # Convert pixel values (as used by Google Sheets) to these character units.
    char_widths = {col: round((px - 5) / 7, 2) for col, px in pixel_widths.items()}

    for col, width in char_widths.items():
        ws.column_dimensions[col].width = width


autofit_columns(ws)
set_fixed_column_widths(ws)

# ─── Step 3: Apply Filters ────────────────────────────────────────────────
filter_cols = ["Plant Type", "Bloom Color", "Sun", "Water", "Attracts"]
header = [cell.value for cell in ws[1]]
filter_indices = [i + 1 for i, val in enumerate(header) if val in filter_cols]
if filter_indices:
    col_range = f"{get_column_letter(min(filter_indices))}1:{get_column_letter(max(filter_indices))}1"
    ws.auto_filter.ref = col_range

# ─── Step 4: Format Cells + Short Hyperlinks ──────────────────────────────
link_map = {
    "Link: Missouri Botanical Garden": "[MBG]",
    "Link: Wildflower.org": "[WF]",
    "Link: Pleasantrunnursery.com": "[PR]",
    "Link: Newmoonnursery.com": "[NM]",
    "Link: Pinelandsnursery.com": "[PN]",
}


def style_sheet(ws: Worksheet, df: pd.DataFrame, header: list[str]) -> None:
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

    for row_idx, row in enumerate(df.itertuples(index=False, name=None), start=2):
        for col_idx, (col_name, value) in enumerate(zip(header, row), start=1):
            cell: Cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = link_map[col_name]
            cell.hyperlink = value
            cell.style = "Hyperlink"
        else:
             cell.value = value

        if col_name == "Botanical Name":
                cell.font = Font(italic=True)
        if not value:
                cell.fill = red_fill


style_sheet(ws, df, header)

# ─── Step 4B: Add raw export sheet with full link data ─────────────────────
raw_sheet = wb.create_sheet("Plant Data CSV — No Short Links")
for col_idx, col_name in enumerate(df.columns, start=1):
    cell = raw_sheet.cell(row=1, column=col_idx)
    cell.value = col_name
    cell.font = BOLD_FONT

for row_idx, row in enumerate(df.itertuples(index=False, name=None), start=2):
    for col_idx, value in enumerate(row, start=1):
        raw_sheet.cell(row=row_idx, column=col_idx).value = value

set_fixed_column_widths(raw_sheet)


# ─── Step 4: Format Cells + Short Hyperlinks ──────────────────────────────
link_map = {
    "Link: Missouri Botanical Garden": "[MBG]",
    "Link: Wildflower.org": "[WF]",
    "Link: Pleasantrunnursery.com": "[PR]",
    "Link: Newmoonnursery.com": "[NM]",
    "Link: Pinelandsnursery.com": "[PN]",
}


def style_sheet(ws: Worksheet, df: pd.DataFrame, header: list[str]) -> None:
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

    for row_idx, row in enumerate(df.itertuples(index=False, name=None), start=2):
        for col_idx, (col_name, value) in enumerate(zip(header, row), start=1):
            cell: Cell = ws.cell(row=row_idx, column=col_idx)
            value = str(value).strip()

            if col_name in link_map and value.startswith("http"):
                cell.value = link_map[col_name]
                cell.hyperlink = value
                cell.style = "Hyperlink"
            else:
                cell.value = value

            if col_name == "Botanical Name":
                cell.font = Font(italic=True)
            if not value:
                cell.fill = red_fill


style_sheet(ws, df, header)

# ─── Step 4B: Add raw export sheet with full link data ─────────────────────
raw_sheet = wb.create_sheet("Plant Data CSV — No Short Links")
for col_idx, col_name in enumerate(df.columns, start=1):
    cell = raw_sheet.cell(row=1, column=col_idx)
    cell.value = col_name
    cell.font = BOLD_FONT

for row_idx, row in enumerate(df.itertuples(index=False, name=None), start=2):
    for col_idx, value in enumerate(row, start=1):
        raw_sheet.cell(row=row_idx, column=col_idx).value = value

set_fixed_column_widths(raw_sheet)

# ─── Step 5: README Sheet ─────────────────────────────────────────────────
readme = wb.create_sheet("README")
readme.sheet_properties.tabColor = "A9A9A9"
readme.column_dimensions["A"].width = 100
readme["A1"] = "🌿 Plant Database Export: Excel Legend and Process Notes"
readme["A3"] = "Legend:"
readme["A4"] = "🔴 Red: Missing value (empty cell)"
readme["A6"] = "Filters applied only to these columns:"
readme["A7"] = ", ".join(filter_cols)
readme["A9"] = "🛠 How to filter by partial match in Excel:"
readme["A10"] = (
    "1. Click the filter dropdown on the column header (e.g., Sun or Characteristics)."
)
readme["A11"] = "2. Choose 'Text Filters' > 'Contains...'"
readme["A12"] = "3. Type a partial term (e.g., 'shade', 'yellow') and click OK."
readme["A13"] = "💡 Use this to find plants matching conditions across categories."
readme["A15"] = "📄 https://github.com/InfusedChooch/Plants"
readme["A16"] = (
    "This Excel was generated from the filled CSV using the script Excelify2.py"
)
readme["A17"] = (
    "Download https://portableapps.com/apps/internet/google_chrome_portable and place it in the /Static folder"
)
readme["A18"] = "Static/GoogleChromePortable"

# ─── Step 6: Script Version Info ──────────────────────────────────────────
script_descriptions = {
    "Static/Python/PDFScraper.py": "Extracts plant data from the PDF guide",
    "Static/Python/GetLinks.py": "Finds official MBG & WF URLs for each plant",
    "Static/Python/FillMissingData.py": "Populates missing fields using those links",
    "Static/Python/GeneratePDF.py": "Creates printable PDF guide with images and sections",
    "Static/Python/Excelify2.py": "Creates formatted Excel output with filters & highlights",
}
row_start = readme.max_row + 2
readme[f"A{row_start}"] = "📁 Script Version Info (Last Modified):"
for i, (script_path, description) in enumerate(
    script_descriptions.items(), start=row_start + 1
):
    full_path = REPO / script_path
    if full_path.exists():
        modified = datetime.fromtimestamp(full_path.stat().st_mtime).strftime(
            "%Y-%m-%d %H:%M"
        )
        readme[f"A{i}"] = f"{script_path:<40} → {modified}    {description}"
    else:
        readme[f"A{i}"] = f"{script_path:<40} → MISSING        {description}"

# ─── Step 7: Add pip requirements ─────────────────────────────────────────
req_path = REPO / "requirements.txt"
readme_row = readme.max_row + 2
readme[f"A{readme_row}"] = "📦 Required Python Packages:"
try:
    lines = req_path.read_text(encoding="utf-8").splitlines()
    for i, line in enumerate(lines, start=readme_row + 1):
        if line.strip() and not line.strip().startswith("#"):
            readme[f"A{i}"] = line.strip()
except Exception as e:
    readme[f"A{readme_row + 1}"] = f"⚠️ Error reading requirements.txt: {e}"

# ─── Step 8: Embed Black-formatted code ──────────────────────────────────
for script_path, description in script_descriptions.items():
    full_path = REPO / script_path
    if not full_path.exists():
        continue

    with open(full_path, "r", encoding="utf-8") as f:
        raw_code = f.read()

    try:
        formatted_code = black.format_str(raw_code, mode=black.Mode())
    except Exception:
        formatted_code = raw_code

    ws_embed = wb.create_sheet(Path(script_path).name)
    ws_embed.column_dimensions["A"].width = 120
    ws_embed["A1"] = "```python"
    for i, line in enumerate(formatted_code.splitlines(), start=2):
        ws_embed[f"A{i}"] = line
    ws_embed[f"A{i+1}"] = "```"

# ─── Step 9: Import README.md ─────────────────────────────────────────────
readme_md_path = REPO / "readme.md"
if readme_md_path.exists():
    readme_full = wb.create_sheet("README_full")
    readme_full.column_dimensions["A"].width = 120
    with readme_md_path.open("r", encoding="utf-8") as f:
        for i, line in enumerate(f, start=1):
            readme_full[f"A{i}"] = line.rstrip("\n")
else:
    print("[WARN] readme.md not found. Skipping README_full tab.")

# ─── Step 10: Save and Done ──────────────────────────────────────────────
wb.save(XLSX_FILE)
print(f"Yeehaw Final Excel saved --> {XLSX_FILE}")
