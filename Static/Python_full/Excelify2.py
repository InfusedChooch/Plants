#!/usr/bin/env python3
# Excelify2.py – Create a styled Excel workbook from the fully populated plant CSV.
# 2025-06-11  · adds source-legend row and NA-link highlighting
# -----------------------------------------------------------------------------

from pathlib import Path
import sys, argparse, pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
import black

# ── CLI ----------------------------------------------------------------------
parser = argparse.ArgumentParser(description="Export formatted Excel from CSV")
parser.add_argument("--in_csv",      default="Outputs/Plants_Linked_Filled.csv",
                    help="Input CSV file with filled data")
parser.add_argument("--out_xlsx",    default="Outputs/Plants_Linked_Filled_Review.xlsx",
                    help="Output Excel file")
parser.add_argument("--template_csv",default="Templates/0611_Masterlist_New_Beta_Nodata.csv",
                    help="CSV that defines final column order")
args = parser.parse_args()

# ── Project-root helper ------------------------------------------------------
def repo_dir() -> Path:
    """Return repo root that contains Templates/ and Outputs/."""
    exe_path = Path(sys.executable if getattr(sys, "frozen", False) else __file__).resolve()
    for parent in exe_path.parents:
        if (parent / "Templates").is_dir() and (parent / "Outputs").is_dir():
            return parent
    return exe_path.parent

REPO         = repo_dir()
CSV_FILE     = (REPO / args.in_csv).resolve()
XLSX_FILE    = (REPO / args.out_xlsx).resolve()
TEMPLATE_CSV = (REPO / args.template_csv).resolve()
XLSX_FILE.parent.mkdir(parents=True, exist_ok=True)

# ── Step 1 · Load CSV, align columns, write base workbook --------------------
df            = pd.read_csv(CSV_FILE, dtype=str, encoding="utf-8-sig", keep_default_na=False).fillna("")
template_cols = list(pd.read_csv(TEMPLATE_CSV, nrows=0, keep_default_na=False).columns)
df            = df.reindex(columns=template_cols + [c for c in df.columns if c not in template_cols])

# normalise casing
if "Common Name" in df.columns:
    df["Common Name"] = df["Common Name"].str.upper()
if "Botanical Name" in df.columns:
    df["Botanical Name"] = df["Botanical Name"].apply(lambda n: " ".join(w.capitalize() for w in str(n).split()))

df.to_excel(XLSX_FILE, index=False)
wb, ws = load_workbook(XLSX_FILE), None
ws = wb.active
ws.title = "Plant Data"

# ── NEW · column-source cheat-sheet -----------------------------------------
DATA_SOURCE: dict[str, str] = {
    "Plant Type":                   "Masterlist",
    "Key":                          "FillMissingData",
    "Botanical Name":               "Masterlist",
    "Common Name":                  "Masterlist",
    "Height (ft)":                  "MBG -> WF -> Pinelands",
    "Spread (ft)":                  "MBG -> WF -> Pinelands",
    "Bloom Color":                  "WF + MBG + Pinelands/NM",
    "Bloom Time":                   "WF + MBG + Pinelands/NM",
    "Sun":                          "MBG -> WF “Light Req.”",
    "Water":                        "MBG -> WF “Soil Moisture”",
    "AGCP Regional Status":         "WF (Wetland Indicator)",
    "USDA Hardiness Zone":          "MBG “Zone”",
    "Attracts":                     "PR + WF + MBG + Pinelands",
    "Tolerates":                    "MBG + PR + NM + Pinelands",
    "Soil Description":             "WF “Soil Description”",
    "Condition Comments":           "WF “Condition Comments”",
    "MaintenanceLevel":             "MBG “Maintenance”",
    "Native Habitats":              "WF “Native Habitat”",
    "Culture":                      "MBG “Culture”",
    "Uses":                         "MBG “Uses”",
    "UseXYZ":                       "WF Benefit list",
    "WFMaintenance":                "WF “Maintenance:”",
    "Problems":                     "MBG “Problems”",
    "Link: Missouri Botanical Garden": "GetLinks (MBG ID)",
    "Link: Wildflower.org":             "GetLinks (USDA ID)",
    "Link: Pleasantrunnursery.com":     "GetLinks (name)",
    "Link: Newmoonnursery.com":         "GetLinks (name)",
    "Link: Pinelandsnursery.com":       "GetLinks (name)",
    "Rev":                              "User Input (YYYYMMDD_FL)",
}

# insert legend row just below headers
ws.insert_rows(2)
for col_idx, col_name in enumerate([c.value for c in ws[1]], start=1):
    cell = ws.cell(row=2, column=col_idx)
    cell.value = DATA_SOURCE.get(col_name, "")
    cell.font  = Font(italic=True, size=8)
    cell.alignment = cell.alignment.copy(wrap_text=True, shrink_to_fit=True)

# ── Step 2 · style header & freeze panes ------------------------------------
HEADER_FILL = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
for c in ws[1]:
    c.fill = HEADER_FILL
    c.font = Font(bold=True)

ws.freeze_panes = "A3"  # header + legend fixed

def autofit(ws: Worksheet):
    for c in ws[1]:
        ws.column_dimensions[get_column_letter(c.col_idx)].width = len(str(c.value or "")) + 2
autofit(ws)

# ── Step 3 · filters on key columns -----------------------------------------
FILTER_COLS = ["Plant Type", "Bloom Color", "Sun", "Water", "Attracts"]
hdr = [c.value for c in ws[1]]
idx = [i+1 for i,v in enumerate(hdr) if v in FILTER_COLS]
if idx:
    ws.auto_filter.ref = f"{get_column_letter(min(idx))}1:{get_column_letter(max(idx))}1"

# ── Step 4 · cell formatting & hyperlinks -----------------------------------
MISSING_FILL = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
NA_LINK_FILL = PatternFill(start_color="B7D7FF", end_color="B7D7FF", fill_type="solid")
REV_FILLED   = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # light green

def style_sheet(ws: Worksheet, df: pd.DataFrame, header: list[str]) -> None:
    for r_idx, row in enumerate(df.itertuples(index=False, name=None), start=3):  # data starts at row 3
        for c_idx, (col_name, val) in enumerate(zip(header, row), start=1):
            cell: Cell = ws.cell(row=r_idx, column=c_idx)
            val = str(val).strip()

            if not val:
                cell.value, cell.fill = "Needs Review", MISSING_FILL
            else:
                cell.value = val

                if col_name.startswith("Link: "):
                    if val == "NA":
                        cell.fill = NA_LINK_FILL
                    elif val.lower().startswith("http"):
                        cell.hyperlink = val
                        cell.style = "Hyperlink"
                elif val.lower().startswith("http"):
                    cell.hyperlink = val
                    cell.style = "Hyperlink"

                # ✅ Highlight filled Rev cells
                if col_name.strip().lower() == "rev":
                    cell.fill = REV_FILLED

                # ✅ Format Botanical Name
                if col_name.strip().lower() == "botanical name":
                    parts = val.split()
                    if len(parts) >= 2:
                        genus = parts[0].capitalize()
                        species = parts[1].lower()
                        variety = " ".join(parts[2:]).strip("' ")
                        if variety:
                            variety = f"'{variety.title()}'"
                        else:
                            variety = ""
                        formatted = " ".join(filter(None, [genus, species, variety]))
                        cell.value = formatted
                    cell.font = Font(italic=True)

style_sheet(ws, df, hdr)


excel_widths = {
    "A": 22,
    "B": 9,
    "C": 26,
    "D": 23,
    "E": 13,
    "G": 13,
    "H": 22,
    "I": 22,
    "J": 22,
    "K": 18,
    "L": 18,
    "M": 43,
    "N": 26,
    "O": 25,
    "P": 25,
    "Q": 22,
    "R": 22,
    "S": 22,
    "T": 22,
    "U": 22,
    "V": 22,
    "W": 22,
    "X": 22,
    "Y": 22,
    "Z": 22,
    "AA": 22,
    "AB": 22,
    "AC": 22,
}

# width tweaks for known long columns
ws.column_dimensions["A"].width  = 22
ws.column_dimensions["C"].width  = 26
ws.column_dimensions["M"].width  = 43
ws.column_dimensions["O"].width  = 25

# ── Step 5 · README sheet ---------------------------------------------------
readme = wb.create_sheet("README")
readme.sheet_properties.tabColor = "A9A9A9"
readme.column_dimensions["A"].width = 100
readme["A1"] = " Plant Database Export: Excel Legend and Process Notes"
readme["A3"] = "Legend:"
readme["A4"] = "RED: Missing value (empty cell)"
readme["A5"] = "BLUE: Link cell still ‘NA’ (needs lookup)"
readme["A7"] = "Filters applied only to these columns:"
readme["A8"] = ", ".join(FILTER_COLS)
readme["A10"] = "Tools: How to filter by partial match in Excel:"
readme["A11"] = "1. Click the filter dropdown on the header."
readme["A12"] = "2. Choose 'Text Filters' ➜ 'Contains...'"
readme["A13"] = "3. Type part of a word (e.g., 'shade', 'yellow')."

# -- Step 6 · embed helper scripts ----------------------------------------
from pathlib import Path
import black

def find_script_root(repo: Path) -> Path:
    """
    Locate the folder that holds the full-version helper scripts.

    1) <repo>/Static/Python_full           ← current layout
    2) <repo>/_internal/Static/Python_full ← legacy layout
    """
    for candidate in (
        repo / "Static" / "Python_full",
        repo / "_internal" / "Static" / "Python_full",
    ):
        if candidate.is_dir():
            return candidate.resolve()
    raise FileNotFoundError("Cannot locate Static/Python_full")

PYTHON_FULL = find_script_root(REPO)

script_descriptions = {
    "PDFScraper.py":      "Extract plant data from source PDF",
    "GetLinks.py":        "Find MBG & WF links",
    "FillMissingData.py": "Populate missing fields",
    "GeneratePDF.py":     "Create formatted PDF guide",
    "Excelify2.py":       "Make this Excel workbook",
}

for script, desc in script_descriptions.items():
    src = PYTHON_FULL / script
    if not src.exists():
        continue                       # skip helpers that are not present
    with src.open(encoding="utf-8") as f:
        raw = f.read()
    try:
        code = black.format_str(raw, mode=black.Mode())
    except Exception:
        code = raw                     # keep original text if Black fails
    ws_code = wb.create_sheet(script)  # one worksheet per helper script
    ws_code.column_dimensions["A"].width = 120
    ws_code["A1"] = f"# {script} - {desc}"
    for i, line in enumerate(code.splitlines(), start=2):
        ws_code[f"A{i}"] = line


# ── Step 7 · pip requirements list -----------------------------------------
req = REPO / "requirements.txt"
row = readme.max_row + 2
readme[f"A{row}"] = "[PKG] Required Python Packages:"
if req.exists():
    for i, line in enumerate(req.read_text().splitlines(), start=row+1):
        if line.strip() and not line.lstrip().startswith("#"):
            readme[f"A{i}"] = line.strip()

# ── Step 8 · pull full README.md (optional) ---------------------------------
readme_md = REPO / "readme.md"
if readme_md.exists():
    tab = wb.create_sheet("README_full")
    tab.column_dimensions["A"].width = 120
    for i, line in enumerate(readme_md.read_text(encoding="utf-8").splitlines(), start=1):
        tab[f"A{i}"] = line

# ── Step 9 · finish ---------------------------------------------------------
wb.save(XLSX_FILE)
print(f"Yeehaw! Workbook saved -> {XLSX_FILE}")
