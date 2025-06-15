#!/usr/bin/env python3
# Excelify2.py – Create a styled Excel workbook from the fully populated plant CSV.
# 2025-06-13 · Adds blanket COLUMN_WIDTHS dict, keep_default_na, and cleans up width logic.
# todo I need to make the Static\Python_full\Excelify2_Testing.py out put resemble the logic in ReviewFiles\Plants_Reformat_Logic.pdf and ReviewFiles\Plants_Reformat_Logic.xlsx.

# todo CSV Format: [Tag1,""URL1"",""Label1""];[Tag2,""URL2"",""Label2""] : "[T1,""https://Test.com"",""Test 1""];[T2,""https://Test.com"",""Test 2""];[T3,""https://Test.com"",""Test 3""]"
# todo New Workbook order: [Readme!][Pretty Data"Plant Data"][Other Links Add/Sub][RAW CSV Export][Imported .py codes][Dir Readme.md]
# todo Excel Format   -- Maybe a Searchable/appendable list 


# ? I need to figure out how to get urls added and edited from the CSV header
# ! This works in excel for Amsonia Hubrichtii	THREADLEAF BLUE STAR	T1	Test 1	https://Test.com	T2	Test 2	https://Test.com	T3	Test 3	https://Test.com
 
# * =TEXTJOIN(";", TRUE,IF(OR($C3="",$E3="",$D3=""), "",CONCAT("[",$C3,",",CHAR(34),$E3,CHAR(34),",",CHAR(34),$D3,CHAR(34),"]")),IF(OR($F3="",$H3="",$G3=""), "",CONCAT("[",$F3,",",CHAR(34),$H3,CHAR(34),",",CHAR(34),$G3,CHAR(34),"]")),IF(OR($I3="",$K3="",$J3=""), "",CONCAT("[",$I3,",",CHAR(34),$K3,CHAR(34),",",CHAR(34),$J3,CHAR(34),"]")))



from pathlib import Path
import sys, argparse, pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.formula.translate import Translator
from openpyxl.workbook.properties import CalcProperties
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.datavalidation import DataValidation
import black 

# ── Column widths (characters) -------------------------------------------
# Tweak values here to resize any column on export.
COLUMN_WIDTHS: dict[str, int] = {
    # core ID fields
    "Plant Type": 20,
    "Key": 7,
    "Botanical Name": 23,
    "Common Name": 23,
    # basics
    "Height (ft)": 10,
    "Spread (ft)": 10,
    "Bloom Color": 18,
    "Bloom Time": 18,
    "Sun": 18,
    "Water": 15,
    "AGCP Regional Status": 18,
    "USDA Hardiness Zone": 16,
    # descriptive
    "Attracts": 24,
    "Tolerates": 28,
    "Soil Description": 32,
    "Condition Comments": 32,
    "MaintenanceLevel": 15,
    "Native Habitats": 26,
    "Culture": 24,
    "Uses": 20,
    "UseXYZ": 20,
    "WFMaintenance": 22,
    "Problems": 25,
    # links
    "Link: Missouri Botanical Garden": 27,
    "Link: Wildflower.org": 27,
    "Link: Pleasantrunnursery.com": 27,
    "Link: Newmoonnursery.com": 27,
    "Link: Pinelandsnursery.com": 27,
    # review
    "Rev": 15,
    "Mark Reviewed": 8,
    # other links
    "Link #":  8,
    "Tag":     6,
    "URL":     43,
    "Label":   18,
}
DEFAULT_WIDTH: int = 18

# ── CLI ------------------------------------------------------------------
parser = argparse.ArgumentParser(description="Export formatted Excel from CSV")
parser.add_argument("--in_csv",      default="Outputs/Excel_Template00.csv")
parser.add_argument("--out_xlsx",    default="ReviewFiles/Plants_01.xlsx")
parser.add_argument("--template_csv",default="Templates/Plants_Template.csv")
args = parser.parse_args()

# ── Paths ----------------------------------------------------------------
# * Find repository root when running from source or bundle
def repo_dir() -> Path:
    exe_path = Path(sys.executable if getattr(sys, "frozen", False) else __file__).resolve()
    for parent in exe_path.parents:
        if (parent / "Templates").is_dir() and (parent / "Outputs").is_dir():
            return parent
    return exe_path.parent

REPO         = repo_dir()
CSV_FILE     = (REPO / args.in_csv).resolve()
XLSX_FILE    = (REPO / args.out_xlsx).resolve()
TEMPLATE_CSV = (REPO / args.template_csv).resolve()
XLSX_FILE.parent.mkdir(parents=True, exist_ok=True)

# ── Load CSV, align columns ----------------------------------------------
df = (
    pd.read_csv(
        CSV_FILE,
        dtype=str,
        encoding="utf-8-sig",
        keep_default_na=False,       # ← keeps literal “NA” strings
    )
    .fillna("")
)
template_cols = list(
    pd.read_csv(TEMPLATE_CSV, nrows=0, keep_default_na=False).columns
)
df = df.reindex(columns=template_cols + [c for c in df.columns if c not in template_cols])

# * Split "Link: Others" into editable columns
def _parse_other_links(text: str) -> list[tuple[str, str, str]]:
    """Return list of (tag, url, label) tuples from the Link: Others cell."""
    import re

    pattern = r"\[(?P<tag>[^,\]]+),\"(?P<url>[^\"]+)\",\"(?P<label>[^\"]+)\"\]"
    return re.findall(pattern, text or "")

# * Helper: build TEXTJOIN formula for Other Links rows
def build_link_formula(row: int, max_links: int) -> list[str]:
    """
    Return list of Excel formulas, one per [Tag, URL, Label] group.
    Used for injecting staged link chunks into helper columns.
    """
    formulas = []
    for idx in range(1, max_links + 1):
        base = 3 * (idx - 1) + 3  # C, F, I...
        tag = get_column_letter(base)
        url = get_column_letter(base + 1)
        lab = get_column_letter(base + 2)

        formula = (
            f'=IF(OR({tag}{row}="",{url}{row}="",{lab}{row}=""),"",'
            f'CONCAT("[",{tag}{row},",",CHAR(34),{url}{row},CHAR(34),",",CHAR(34),{lab}{row},CHAR(34),"]"))'
        )
        formulas.append(formula)
    return formulas


if "Link: Others" in df.columns:
    links_parsed = df["Link: Others"].apply(_parse_other_links)
    max_links = links_parsed.map(len).max()

    for idx in range(max_links):
        df[f"Other Tag {idx+1}"] = links_parsed.apply(
            lambda lst: lst[idx][0] if idx < len(lst) else ""
        )
        df[f"Other URL {idx+1}"] = links_parsed.apply(
            lambda lst: lst[idx][1] if idx < len(lst) else ""
        )
        df[f"Other Label {idx+1}"] = links_parsed.apply(
            lambda lst: lst[idx][2] if idx < len(lst) else ""
        )

        
MAX_LINKS = locals().get("max_links", 5)

# allow up to N links per record (makes "1,2,3,..." for validation lists)
LINK_CHOICES = ",".join(str(i) for i in range(1, MAX_LINKS + 1))  # e.g. "1,2,3"


# Normalise casing
if "Common Name" in df.columns:
    df["Common Name"] = df["Common Name"].str.upper()
if "Botanical Name" in df.columns:
    df["Botanical Name"] = df["Botanical Name"].apply(
        lambda n: " ".join(w.capitalize() for w in str(n).split())
    )

# ── Inject “Mark Reviewed” column if missing -----------------------------
if "Mark Reviewed" not in df.columns:
    df.insert(df.columns.get_loc("Rev") + 1, "Mark Reviewed", "")

# * track sizes for later formulas
DATA_ROWS = len(df)
MAX_LINKS = locals().get("max_links", 5)

# ── Canonical column lists ────────────────────────────────────────────
LINK_LIST_COLS = ["Link #", "Tag", "URL", "Label"]
MASTER_COLS = [
    "Plant Type","Key","Botanical Name","Common Name",
    "Height (ft)","Spread (ft)","Bloom Color","Bloom Time",
    "Sun","Water","AGCP Regional Status","USDA Hardiness Zone",
    "Attracts","Tolerates","Soil Description","Condition Comments",
    "MaintenanceLevel","Native Habitats","Culture","Uses",
    "UseXYZ","WFMaintenance","Problems",
    "Link: Missouri Botanical Garden","Link: Wildflower.org",
    "Link: Pleasantrunnursery.com","Link: Newmoonnursery.com",
    "Link: Pinelandsnursery.com","Link: Others","Rev",
]

# Plant Data = master columns + Mark Reviewed (no extra link columns)
PLANT_DATA_COLS = MASTER_COLS + ["Mark Reviewed"] + LINK_LIST_COLS
DISPLAY_PLANT_DATA_COLS = [c for c in PLANT_DATA_COLS if c != "Link: Others"]
for col in LINK_LIST_COLS:
    if col not in df.columns:
        df[col] = ""                          # blank until Excel formulas fill


# ── Build the display-only DataFrame & write it to disk -------------------
DISPLAY_PLANT_DATA_COLS = [c for c in PLANT_DATA_COLS if c != "Link: Others"]
df_out = df.reindex(columns=DISPLAY_PLANT_DATA_COLS, fill_value="")
PLANT_DATA_HEADERS = DISPLAY_PLANT_DATA_COLS

# IMPORTANT: actually create / overwrite the workbook *before* customising it
df_out.to_excel(XLSX_FILE, index=False, na_rep="NA", sheet_name="Plant Data")

# ── Load the fresh workbook and start styling ----------------------------
from openpyxl import load_workbook
wb = load_workbook(XLSX_FILE)
ws = wb.active
ws.title = "Plant Data"

# ── RAW CSV Export ────────────────────────────────────────────────────

raw_ws = wb.create_sheet("RAW CSV Export")
raw_ws.append(MASTER_COLS)                     # header row

for r in range(DATA_ROWS):                    # 1-based to match Excel rows
    for c_idx, col_name in enumerate(MASTER_COLS, start=1):
        col_letter = get_column_letter(c_idx)

        formula_col = get_column_letter(3 * MAX_LINKS + 3)   # 3·5+3 = 18 → "R"

        # ‘Link: Others’ comes from the helper sheet; everything else mirrors Plant Data
        if col_name == "Link: Others":
            raw_ws.cell(row=r+2, column=c_idx).value = (
                f"='Other Links'!${formula_col}{r+3}"
            )
        else:
            # Plant Data has headers in row 2 and the dataset starts on row 3
            raw_ws.cell(row=r+2, column=c_idx).value = (
                f"='Plant Data'!{col_letter}{r+2}"
            )

# ── Source-legend row  ----------------------------------------
DATA_SOURCE = {
    "Plant Type": "Masterlist",
    "Key": "FillMissingData",
    "Botanical Name": "Masterlist",
    "Common Name": "Masterlist",
    "Height (ft)": "MBG -> WF -> Pinelands",
    "Spread (ft)": "MBG -> WF -> Pinelands",
    "Bloom Color": "WF + MBG + Pinelands/NM",
    "Bloom Time": "WF + MBG + Pinelands/NM",
    "Sun": "MBG -> WF “Light Req.”",
    "Water": "MBG -> WF “Soil Moisture”",
    "AGCP Regional Status": "WF (Wetland Indicator)",
    "USDA Hardiness Zone": "MBG “Zone”",
    "Attracts": "PR + WF + MBG + Pinelands",
    "Tolerates": "MBG + PR + NM + Pinelands",
    "Soil Description": "WF “Soil Description”",
    "Condition Comments": "WF “Condition Comments”",
    "MaintenanceLevel": "MBG “Maintenance”",
    "Native Habitats": "WF “Native Habitat”",
    "Culture": "MBG “Culture”",
    "Uses": "MBG “Uses”",
    "UseXYZ": "WF Benefit list",
    "WFMaintenance": "WF “Maintenance:”",
    "Problems": "MBG “Problems”",
    "Link: Missouri Botanical Garden": "GetLinks (MBG ID)",
    "Link: Wildflower.org": "GetLinks (USDA ID)",
    "Link: Pleasantrunnursery.com": "GetLinks (name)",
    "Link: Newmoonnursery.com": "GetLinks (name)",
    "Link: Pinelandsnursery.com": "GetLinks (name)",
    "Rev": "User Input (YYYYMMDD_FL)",
    "Mark Reviewed": "Type Initials",
    "Link: Others": "Edit on `Other Links`",
    "Link Lists": "Populated from `Other Links`",
    
}


# ── Merge helper columns under single 'Link Lists' header ─────────────
col_first = PLANT_DATA_HEADERS.index("Link #") + 1
col_last  = col_first + 3
ws.merge_cells(start_row=1, start_column=col_first, end_row=1, end_column=col_last)
cell = ws.cell(row=1, column=col_first)
cell.value = "Link Lists"
cell.font = Font(bold=True, size=11)
cell.alignment = Alignment(horizontal="center", vertical="center")

ws.insert_rows(2)
for col_idx, col_name in enumerate([c.value for c in ws[1]], start=1):
    cell = ws.cell(row=2, column=col_idx)
    cell.value = DATA_SOURCE.get(col_name, "")
    cell.font = Font(italic=True, size=8)
    cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=False, shrink_to_fit=False)

# ── Header formatting + freeze -------------------------------------------
HEADER_FILL = PatternFill(start_color="CFE2F3", end_color="CFE2F3", fill_type="solid")
for c in ws[1]:
    c.fill = HEADER_FILL
    c.font = Font(bold=True, size=11)
    c.alignment = Alignment(horizontal="center", vertical="center")
ws.freeze_panes = "E3"

# ── Autofit helper --------------------------------------------------------
# * Resize columns to fit content while respecting caps
def autofit(ws: Worksheet) -> None:
    long_fields = {
        "UseXYZ": 50, "Culture": 50, "Uses": 50,
        "Soil Description": 48, "Condition Comments": 48, "Native Habitats": 42,
    }
    headers = [c.value for c in ws[1]]
    for col_cells in ws.iter_cols(min_row=1, max_row=ws.max_row):
        col_idx = col_cells[0].column
        col_letter = get_column_letter(col_idx)
        header = headers[col_idx - 1] if col_idx - 1 < len(headers) else ""
        lengths = [
            len(str(cell.value or ""))
            for cell in col_cells
            if str(cell.value or "").strip() not in {"Needs Review", ""}
        ]
        if lengths:
            cap = long_fields.get(header, 64)
            ws.column_dimensions[col_letter].width = min(max(lengths) + 2, cap)

# ── Style all data cells (unchanged) -------------------------------------
MISSING_FILL = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
NA_LINK_FILL = PatternFill(start_color="B7D7FF", end_color="B7D7FF", fill_type="solid")
REV_FILLED   = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

# * Apply row-level styling and hyperlink logic
def style_sheet(ws: Worksheet, df: pd.DataFrame, header: list[str]) -> None:
    REV_MISSING_FILL = PatternFill(start_color="FFF79A", end_color="FFF79A", fill_type="solid")
    ROW_ALT_FILL = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")
    long_wrap_fields = {
        # already here:
        "UseXYZ", "Culture", "Uses",
        "Soil Description", "Condition Comments", "Native Habitats",
        "Rev",
        # add the two list-heavy ones:
        "Attracts", "Tolerates",
    }


    for r_idx, row in enumerate(df.itertuples(index=False, name=None), start=3):
        for c_idx, (col_name, val) in enumerate(zip(header, row), start=1):
            cell = ws.cell(row=r_idx, column=c_idx)
            val = str(val).strip()
            # Helper columns Tag/URL/Label skipped until formulas
            if col_name in {"Tag", "URL", "Label"}:
                cell.value = ""
                cell.alignment = Alignment(horizontal="center", vertical="top")
                continue
            col_name_lower = col_name.strip().lower()

            # Rev column handling
            if col_name_lower == "rev":
                if val:
                    cell.value = val
                    cell.fill = REV_FILLED
                else:
                    mark_col_letter = get_column_letter(c_idx + 1)
                    cell.value = (
                        f'=IF({mark_col_letter}{r_idx}<>"",'
                        f' TEXT(TODAY(),"yyyymmdd") & "_" & {mark_col_letter}{r_idx}, "")'
                    )
                    cell.fill = REV_MISSING_FILL
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                continue

            # Mark Reviewed column
            if col_name_lower == "mark reviewed":
                cell.value = val if val else ""
                cell.alignment = Alignment(horizontal="center", vertical="center")
                continue

            # Links – NA / Needs Review logic
            if val.upper() == "NA" and col_name.startswith("Link: "):
                cell.value, cell.fill = "NA", NA_LINK_FILL
            elif not val:
                cell.value, cell.fill = "Needs Review", MISSING_FILL
            else:
                cell.value = val
                if col_name.startswith("Link: ") and val.lower().startswith("http"):
                    cell.hyperlink = val
                    cell.style = "Hyperlink"
                    cell.font = Font(color="0000EE", underline="single")

                # Italicise Botanical Name
                if col_name_lower == "botanical name":
                    parts = val.split()
                    if len(parts) >= 2:
                        genus, species = parts[0].capitalize(), parts[1].lower()
                        variety = " ".join(parts[2:]).strip("' ")
                        variety = f"'{variety.title()}'" if variety else ""
                        cell.value = " ".join(filter(None, [genus, species, variety]))
                    cell.font = Font(italic=True)

            # Text wrapping
            if col_name in long_wrap_fields:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            else:
                cell.alignment = Alignment(wrap_text=False, vertical="top")

        # alt-row shading
        if r_idx % 2 == 1:
            for cur in ws[r_idx]:
                if cur.fill == PatternFill():
                    cur.fill = ROW_ALT_FILL
        
                # ── auto-grow row height when wrapping is on --------------------
        max_lines = 1
        for cur in ws[r_idx]:
            if cur.alignment and cur.alignment.wrap_text:
                # a) explicit line-breaks             → count '\n'
                # b) long text that Excel will wrap   → rough 50-char estimate
                lines = max(
                    str(cur.value).count("\n") + 1,
                    len(str(cur.value or "")) // 50 + 1,
                )
                max_lines = max(max_lines, lines)

        if max_lines > 1:
            ws.row_dimensions[r_idx].height = 15 * max_lines  # ≈15 px per line


# ── Style + Fit ----------------------------------------------------------
style_sheet(ws, df_out, PLANT_DATA_HEADERS)
autofit(ws)

# ── Link-list helper columns (AD … AG) -----------------------------------
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

LINK_DV   = ",".join(f"Link {n}" for n in range(1, MAX_LINKS + 1))  # "Link 1,Link 2,…"
col_LinkN = PLANT_DATA_HEADERS.index("Link #") + 1  # AD → 1-based index
col_Tag   = col_LinkN + 1                           # AE
col_URL   = col_LinkN + 2                           # AF
col_Label = col_LinkN + 3                           # AG

for r in range(3, ws.max_row + 1):                  # data start at row 3
    # --- drop-down in “Link #” ------------------------------------------------
    dv = DataValidation(type="list", formula1=f'"{LINK_DV}"', allow_blank=False)
    ws.add_data_validation(dv)
    dv.add(ws.cell(row=r, column=col_LinkN))
    ws.cell(row=r, column=col_LinkN).value = "Link 1"          # sensible default

# ---- build CHOOSE() argument list (Tag / URL / Label) compatible with all Excel versions ----
for r in range(3, ws.max_row + 1):
    args_tag, args_url, args_label = [], [], []
    for n in range(1, MAX_LINKS + 1):
        base = 3 * (n - 1) + 3
        tag  = get_column_letter(base)
        url  = get_column_letter(base + 1)
        lab  = get_column_letter(base + 2)
        args_tag.append(f"'Other Links'!${tag}${r}")
        args_url.append(f"'Other Links'!${url}${r}")
        args_label.append(f"'Other Links'!${lab}${r}")

    choice_idx = f'MATCH(${get_column_letter(col_LinkN)}{r},{{"Link 1","Link 2","Link 3","Link 4","Link 5"}},0)'

    ws.cell(row=r, column=col_Tag).value   = f"=CHOOSE({choice_idx},{','.join(args_tag)})"
    ws.cell(row=r, column=col_URL).value   = f"=CHOOSE({choice_idx},{','.join(args_url)})"
    ws.cell(row=r, column=col_Label).value = f"=CHOOSE({choice_idx},{','.join(args_label)})"
    ws.cell(row=r, column=col_URL).style   = "Hyperlink"

    


for hdr_cell in ws[1]:
    hdr = str(hdr_cell.value).strip()
    col_letter = get_column_letter(hdr_cell.column)
    ws.column_dimensions[col_letter].width = COLUMN_WIDTHS.get(hdr, DEFAULT_WIDTH)

# Row heights
for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
    ws.row_dimensions[row[0].row].height = 28
# ── Other Links sheet (two-row header, Link 1-4 groups) ──────────────────
other_ws = wb.create_sheet("Other Links")

# --- Row 1 · group titles -------------------------------------------------
row1 = ["Botanical Name", "Common Name"]
for i in range(1, MAX_LINKS + 1):             # Link 1 … Link N
    row1 += [f"Link {i}", "", ""]             # merged cells (title spans 3)
row1 += ["Formula", "CSV Imported"]  # formula + raw CSV import
other_ws.append(row1)

# Add 'Match Status' header (merged over row 1 and 2)
match_col_idx = len(row1) + 1
match_col_letter = get_column_letter(match_col_idx)

other_ws.cell(row=1, column=match_col_idx).value = "Match Status"
other_ws.merge_cells(start_row=1, start_column=match_col_idx, end_row=2, end_column=match_col_idx)
other_ws[f"{match_col_letter}1"].font = Font(bold=True)
other_ws[f"{match_col_letter}1"].alignment = Alignment(horizontal="center")
COLUMN_WIDTHS["Match Status"] = 45


# --- Row 2 · sub-titles ---------------------------------------------------
row2 = ["", ""]
for _ in range(MAX_LINKS):
    row2 += ["Tag", "URL", "Label"]
row2 += ["", ""]

# --- merge the Link n title cells ----------------------------------------
col_base = 3
for i in range(MAX_LINKS):
    start = col_base + i * 3
    end   = start + 2
    other_ws.merge_cells(
        start_row=1, start_column=start,
        end_row=1,   end_column=end
    )
# --- bold / center the headers -------------------------------------------
for cell in other_ws[1] + other_ws[2]:
    cell.font      = Font(bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")

# --- append data rows -----------------------------------------------------
for i in range(DATA_ROWS):
    formula_row = other_ws.max_row + 1

    row = [
        df.get("Botanical Name")[i] if "Botanical Name" in df.columns else "",
        df.get("Common Name")[i]    if "Common Name"    in df.columns else "",
    ]
    for idx in range(1, MAX_LINKS + 1):
        row += [
            df.get(f"Other Tag {idx}", [""]*DATA_ROWS)[i],
            df.get(f"Other URL {idx}", [""]*DATA_ROWS)[i],
            df.get(f"Other Label {idx}", [""]*DATA_ROWS)[i],
        ]

    # 1. CSV RAW string
    raw_csv = df.get("Link: Others", [""] * DATA_ROWS)[i]

    # 2. Append dummy cells (we'll fix them after)
    row += ["", raw_csv]  # placeholder for formula + raw
    other_ws.append(row)

    # 3. Inject formula as actual Excel formula
    formula_cell = other_ws.cell(row=formula_row, column=len(row) - 1)
    formula_cell = other_ws.cell(row=formula_row, column=len(row) - 1)
    formula_cell.value = (
        f'=TEXTJOIN(";", TRUE,'
        f'IF(OR(C{formula_row}="",D{formula_row}="",E{formula_row}=""), "",'
        f'CONCAT("[",C{formula_row},",",CHAR(34),D{formula_row},CHAR(34),",",CHAR(34),E{formula_row},CHAR(34),"]")),'
        f'IF(OR(F{formula_row}="",G{formula_row}="",H{formula_row}=""), "",'
        f'CONCAT("[",F{formula_row},",",CHAR(34),G{formula_row},CHAR(34),",",CHAR(34),H{formula_row},CHAR(34),"]")),'
        f'IF(OR(I{formula_row}="",J{formula_row}="",K{formula_row}=""), "",'
        f'CONCAT("[",I{formula_row},",",CHAR(34),J{formula_row},CHAR(34),",",CHAR(34),K{formula_row},CHAR(34),"]")),'
        f'IF(OR(L{formula_row}="",M{formula_row}="",N{formula_row}=""), "",'
        f'CONCAT("[",L{formula_row},",",CHAR(34),M{formula_row},CHAR(34),",",CHAR(34),N{formula_row},CHAR(34),"]")),'
        f'IF(OR(O{formula_row}="",P{formula_row}="",Q{formula_row}=""), "",'
        f'CONCAT("[",O{formula_row},",",CHAR(34),P{formula_row},CHAR(34),",",CHAR(34),Q{formula_row},CHAR(34),"]"))'
        f')'
    )


    if isinstance(formula_cell.value, str) and formula_cell.value.startswith("="):
        formula_cell.value = formula_cell.value  # Nudge Excel to register it
    if isinstance(formula_cell.value, str) and formula_cell.value.startswith("="):
        formula_cell.value = formula_cell.value


    match_col_letter = get_column_letter(len(row))   # CSV RAW
    status_col_letter = get_column_letter(len(row)+1)
    formula_col_letter = get_column_letter(len(row)-1)

    other_ws[f"{status_col_letter}{formula_row}"] = (
        f'=IF({formula_col_letter}{formula_row}={match_col_letter}{formula_row},'
        f'"Matched!",'
        f'CONCAT("Diff → ", {formula_col_letter}{formula_row}, " ≠ ", {match_col_letter}{formula_row}))'
    )



# Get column letters for formula + raw string
col_formula = get_column_letter(len(row) - 1)  # Formula column
col_raw     = get_column_letter(len(row))      # CSV RAW OUTPUT column
start_row   = 3
end_row     = DATA_ROWS + 2

# Excel formula: compare formula cell with raw cell in same row
match_rule = FormulaRule(
    formula=[f'${col_formula}3=${col_raw}3'],
    fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
)
mismatch_rule = FormulaRule(
    formula=[f'${col_formula}3<>${col_raw}3'],
    fill=PatternFill(start_color="FFBABA", end_color="FFBABA", fill_type="solid")
)

# Apply formatting rules to CSV RAW OUTPUT column
other_ws.conditional_formatting.add(f"{col_raw}{start_row}:{col_raw}{end_row}", match_rule)
other_ws.conditional_formatting.add(f"{col_raw}{start_row}:{col_raw}{end_row}", mismatch_rule)




bot_range = f"'Plant Data'!$C$3:$C${DATA_ROWS + 2}"
com_range = f"'Plant Data'!$D$3:$D${DATA_ROWS + 2}"

dv_bot = DataValidation(type="list", formula1=bot_range, allow_blank=True)
dv_com = DataValidation(type="list", formula1=com_range, allow_blank=True)
other_ws.add_data_validation(dv_bot)
other_ws.add_data_validation(dv_com)
dv_bot.add(f"A2:A{DATA_ROWS + 1}")
dv_com.add(f"B2:B{DATA_ROWS + 1}")

# * Link: Others columns in Plant Data reference the formula column

print("[DEBUG] Sample formula preview:")
for i in range(min(3, DATA_ROWS)):
    print(build_link_formula(i + 2, MAX_LINKS))

# ── README -------------------------------------------------------------------
readme = wb.create_sheet("README")
readme.sheet_properties.tabColor = "A9A9A9"
readme.column_dimensions["A"].width = 100

readme["A1"] = "Instructions:"
readme["A3"] = "To record a review:"
readme["A4"] = "1. Type your initials (e.g., AN) in the 'Mark Reviewed' column."
readme["A5"] = "2. The 'Rev' column will auto-fill with today's date and your initials."
readme["A7"] = "Legend:"
readme["A8"] = "RED: Missing value (empty cell)"
readme["A9"] = "BLUE: Link doesn't exist and was reviewd. ‘NA’ =/= No Value"

readme["A11"] = "Filters applied only to these columns:"
readme["A12"] = ", ".join(["Plant Type", "Bloom Color", "Sun", "Water", "Attracts"])

readme["A14"] = "Tools: How to filter by partial match in Excel:"
readme["A15"] = "1. Click the filter dropdown on the header."
readme["A16"] = "2. Choose 'Text Filters' ➜ 'Contains...'"
readme["A17"] = "3. Type part of a word (e.g., 'shade', 'yellow')."
readme["A19"] = "The 'RAW CSV Export' sheet mirrors Plant Data."


# -- Step 6 · embed helper scripts ----------------------------------------
from pathlib import Path
import black

# * Locate Static/Python_full no matter the install layout
def find_script_root(repo: Path) -> Path:
    """
    Locate the folder that holds the full-version helper scripts.

    1) <repo>/Static/Python_full           ← current layout
    2) <repo>/_internal/Static/Python_full ← legacy layout
    """
    for candidate in (
        repo / "Static" / "Python_full",
        repo / "_internal" / "Static" / "Python_full",
    ):
        if candidate.is_dir():
            return candidate.resolve()
    raise FileNotFoundError("Cannot locate Static/Python_full")

PYTHON_FULL = find_script_root(REPO)

script_descriptions = {
    "PDFScraper.py":      "Extract plant data from source PDF",
    "GetLinks.py":        "Find MBG & WF links",
    "FillMissingData.py": "Populate missing fields",
    "GeneratePDF.py":     "Create formatted PDF guide",
    "Excelify2.py":       "Make this Excel workbook",
}

code_sheets: list[Worksheet] = []
for script, desc in script_descriptions.items():
    src = PYTHON_FULL / script
    if not src.exists():
        continue                       # skip helpers that are not present
    with src.open(encoding="utf-8") as f:
        raw = f.read()
    try:
        code = black.format_str(raw, mode=black.Mode())
    except Exception:
        code = raw                     # keep original text if Black fails
    ws_code = wb.create_sheet(script)  # one worksheet per helper script
    code_sheets.append(ws_code)
    ws_code.column_dimensions["A"].width = 120
    ws_code["A1"] = f"# {script} - {desc}"
    for i, line in enumerate(code.splitlines(), start=2):
        ws_code[f"A{i}"] = line


# ── Step 7 · pip requirements list -----------------------------------------
req = REPO / "requirements.txt"
row = readme.max_row + 2
readme[f"A{row}"] = "[PKG] Required Python Packages:"
if req.exists():
    for i, line in enumerate(req.read_text().splitlines(), start=row+1):
        if line.strip() and not line.lstrip().startswith("#"):
            readme[f"A{i}"] = line.strip()

# ── Step 8 · pull full README.md (optional) ---------------------------------
readme_md = REPO / "readme.md"
dir_readme = None
if readme_md.exists():
    dir_readme = wb.create_sheet("Dir README")
    dir_readme.column_dimensions["A"].width = 120
    for i, line in enumerate(readme_md.read_text(encoding="utf-8").splitlines(), start=1):
        dir_readme[f"A{i}"] = line

# * Reorder worksheets per README instructions
sheet_order = [readme, ws, other_ws, raw_ws] + code_sheets
if dir_readme:
    sheet_order.append(dir_readme)
wb._sheets = sheet_order



calc_props = CalcProperties()
calc_props.calcMode        = "auto"   # turn Auto back on
calc_props.fullCalcOnLoad  = True     # recalc on every open
calc_props.forceFullCalc   = True     # for older Excel builds
wb.calculation_properties  = calc_props

# ── Step 9 · finish ---------------------------------------------------------
wb.save(XLSX_FILE)
print(f"Yeehaw! Workbook saved -> {XLSX_FILE}")