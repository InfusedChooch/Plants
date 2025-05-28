# /plants/Excelify.py

from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter

BASE      = Path(__file__).resolve().parent
CSV_FILE  = BASE / "Plants_COMPLETE.csv"
XLSX_FILE = CSV_FILE.with_suffix(".xlsx")

# ─── Load DataFrame and write to Excel ────────────────────────────────────
df = pd.read_csv(CSV_FILE, dtype=str).fillna("")
df.to_excel(XLSX_FILE, index=False)
wb = load_workbook(XLSX_FILE)
ws = wb.active
ws.title = "Plant Data"

# ─── Freeze Panes, Style Headers, Auto Width ──────────────────────────────
ws.freeze_panes = "A2"
HEADER_FILL = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
BOLD_FONT   = Font(bold=True)
for cell in ws[1]:
    cell.fill = HEADER_FILL
    cell.font = BOLD_FONT
for i, column_cells in enumerate(ws.columns, 1):
    max_length = max(len(str(cell.value or "")) for cell in column_cells)
    ws.column_dimensions[get_column_letter(i)].width = min(max_length + 2, 50)
ws.auto_filter.ref = ws.dimensions
ws.sheet_properties.tabColor = "92D050"

# ─── Cell Highlighting Rules ──────────────────────────────────────────────
BLUE   = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
YELLOW = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
RED    = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
mbg_fields = {"Height (ft)", "Spread (ft)", "Sun", "Water", "Characteristics", "Wildlife Benefits", "Distribution"}
wf_fields  = {"Bloom Color", "Bloom Time", "Habitats", "Sun", "Water", "Characteristics", "Wildlife Benefits"}
header     = [cell.value for cell in ws[1]]

for row_idx, row in enumerate(df.itertuples(index=False), start=2):
    mbg_used = row._asdict().get("Link: Missouri Botanical Garden", "").strip().startswith("http")
    wf_used  = row._asdict().get("Link: Wildflower.org", "").strip().startswith("http")
    for col_idx, col_name in enumerate(header, start=1):
        val = getattr(row, col_name.replace(" ", "_"), "")
        cell = ws.cell(row=row_idx, column=col_idx)
        if not cell.value or str(cell.value).strip() == "":
            cell.fill = RED
        elif mbg_used and col_name in mbg_fields:
            cell.fill = BLUE
        elif wf_used and col_name in wf_fields:
            cell.fill = YELLOW

# ─── README Sheet with Color Legend and Code Boxes ────────────────────────
readme = wb.create_sheet("README")
readme.sheet_properties.tabColor = "A9A9A9"
readme.column_dimensions["A"].width = 100

# ─── Text Rows ────────────────────────────────────────────────────────────
readme["A1"] = "🌿 Plant Database Export: Excel Legend and Process Notes"
readme["A3"] = "Legend:"
readme["A4"] = "🔵 Blue: Value filled from Missouri Botanical Garden (MBG)"
readme["A5"] = "🟡 Yellow: Value filled from Wildflower.org (WF)"
readme["A6"] = "🔴 Red: Missing value (empty cell)"

readme["A8"]  = "Workflow:"
readme["A9"]  = "1. PDFScrape.py → Plants_Nolinks.csv"
readme["A10"] = "2. GetLinks_FullHybrid.py → Plants and Links TEST.csv"
readme["A11"] = "3. fill.py → Plants_COMPLETE.csv"
readme["A12"] = "4. Excelify.py → Plants_COMPLETE.xlsx"

# ─── Bordered Code Block Style ────────────────────────────────────────────
gray_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
mono_font = Font(name="Consolas", size=10)
box_border = Border(
    left=Side(border_style="thin", color="AAAAAA"),
    right=Side(border_style="thin", color="AAAAAA"),
    top=Side(border_style="thin", color="AAAAAA"),
    bottom=Side(border_style="thin", color="AAAAAA"),
)

def draw_code_box(start_row: int, title: str):
    readme[f"A{start_row}"] = f"📄 {title}"
    box_rows = range(start_row + 1, start_row + 11)  # 10-row box
    for r in box_rows:
        cell = readme[f"A{r}"]
        cell.fill = gray_fill
        cell.font = mono_font
        cell.border = box_border
        readme.row_dimensions[r].height = 18  # Slightly taller

draw_code_box(15, "PDFScrape.py")
draw_code_box(27, "GetLinks_FullHybrid.py")
draw_code_box(39, "fill.py")
draw_code_box(51, "Excelify.py")

# ─── Save Excel File ──────────────────────────────────────────────────────
wb.save(XLSX_FILE)
print(f"📘 Final Excel saved → {XLSX_FILE}")
